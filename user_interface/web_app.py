import streamlit as st
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import numbers
from openpyxl.styles import Font
from openpyxl.styles import Alignment

import pandas as pd
import io
st.set_page_config(
    page_title="Web app",
    page_icon="🌍",
    layout="centered",  # wide
    initial_sidebar_state="auto")

#Title of the app



# Set up the Streamlit app
st.title('THIS WEB APP FORMATS RAW MARKET SHARE FILES')

st.header("TRY IT!")


# Add a file uploader widget
uploaded_file = st.file_uploader('Upload an Excel file', type=['xlsx'])

year_start_value = st.number_input('Enter first year of the table', value=0, step=1, format='%d')

nb_years_value = st.number_input('Enter number of years featured in the table', value=0, step=1, format='%d')


title_tab=st.text_input('Enter the title of the tab')

header_color = st.text_input('Enter the header color (e.g. FF0000 for red)')

unit_text = st.text_input('Enter the label for the unit column of each year')

ms_text = st.text_input('Enter the label for the market share column of each year')

source_text = st.text_input('Enter the source of the data')

note_text = st.text_input('Enter text for the note section')

def ms_table_creation_web(uploaded_file, year_start_value, nb_years_value, title_tab, header_color, unit_text, ms_text, source_text, note_text):

    df = pd.read_excel(uploaded_file)

    total_row = df.sum(axis=0)
    total_row_df = pd.DataFrame(total_row).T
    total_row_df.loc[0,"Company"]="Total"
    df_complete = df.append(total_row_df, ignore_index=True)


    year_start = year_start_value
    nb_years = nb_years_value
    unit_label = unit_text
    ms_label = ms_text


     # Create a new workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    worksheet = workbook.active


    # Write the DataFrame to the worksheet as a table
    for r in dataframe_to_rows(df_complete, index=False, header=True):
        worksheet.append(r)
    table = openpyxl.worksheet.table.Table(displayName="Table1", ref=f"A1:C{len(df)+1}")
    #table.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    row_between = 3

    # Insert 3 rows at the top of the worksheet
    worksheet.insert_rows(1, amount=row_between)

    # Insert a column at the left of the worksheet
    worksheet.insert_cols(1)

    # Insert a cell with text at the top of the worksheet
    cell = worksheet.cell(row=1, column=1)
    cell.value = title_tab
    cell.font = Font(bold=True)


    # Identify cells for the header of the table
    cell1_header = worksheet.cell(row=(row_between+1), column=2)

    cell2_header = worksheet.cell(row=(row_between+1), column=(2+2*nb_years))

    #Identify cells for the table cell-range

    cell_lower_right = worksheet.cell(row=(1+row_between+df_complete.shape[0]), column=(2+2*nb_years))

    # Percentage format for the MS columns

    j = 2

    first_row = row_between+2
    last_row = row_between+2+df_complete.shape[0]

    for j in range(2,int(2+nb_years)):

        for i in range(first_row, last_row):
            cell_ms = worksheet.cell(row=i, column=2*j)
            cell_ms.number_format = '0%' # possible to have "0.00%" for 2 decimals


    # creating a row with the years

    cell1_header = worksheet.cell(row=(row_between+1), column=2)

    value_stored = cell1_header.value

    cell1_header_above = worksheet.cell(row=(row_between), column=2)

    worksheet.merge_cells(start_row= cell1_header_above.row,
                        start_column= cell1_header_above.column,
                        end_row= cell1_header.row,
                        end_column= cell1_header.column)

    cell1_header_above.value = value_stored

    list_years = [year for year in range(int(year_start), int(year_start + nb_years + 1) )]

    for i,j in zip(range(3,int(2+2*nb_years),2),list_years):
        worksheet.merge_cells(start_row= row_between,
                        start_column= i,
                        end_row= row_between,
                        end_column= i+1)

        cell =  worksheet.cell((row_between), i)
        cell.value =  j
        cell.alignment = Alignment(horizontal='center')

    cell2_header_above = worksheet.cell(row=(row_between), column=(2+2*nb_years))

    for i,j in zip(range(3,int(3+2*nb_years),2),list_years):

        cell = worksheet.cell(row=(row_between+1), column=i)

        cell.value= unit_label

    for i,j in zip(range(4,int(3+2*nb_years),2),list_years):

        cell = worksheet.cell(row=(row_between+1), column=i)

        cell.value= ms_label


    # Define the border style
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    cell_range = worksheet[cell1_header_above.coordinate : cell_lower_right.coordinate]
    for row in cell_range:
        for cell in row:
            cell.border = border



    # Set the background color for the header row of the table
    fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    cell_range = worksheet[cell1_header_above.coordinate : cell2_header.coordinate]
    for row in cell_range:
        for cell in row:
            cell.fill = fill
            cell.font = Font(bold=True)


    # Put in bold last row of the table

    for i in range(2,int(2+2*nb_years+1)):
        cell = worksheet.cell(row=(1+row_between+df_complete.shape[0]), column=i)
        cell.font = Font(bold=True)


    # Adjust column widths to fit their content
    for column in worksheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length
        worksheet.column_dimensions[column_name].width = adjusted_width


    # Source

    cell_source = worksheet.cell(row=(1+row_between+df_complete.shape[0]+2), column=2)
    cell_source.value = source_text

    # Note

    cell_note = worksheet.cell(row=(1+row_between+df_complete.shape[0]+3), column=2)
    cell_note.value = note_text


    # Filter button is on by default - turning it off
    table.autoFilter= None

    with io.BytesIO() as buffer:
        workbook.save(buffer)
        output_file = buffer.getvalue()


   # Save the output workbook to a BytesIO object
    #output_file = openpyxl.writer.excel.save_virtual_workbook()

    # Return the BytesIO object
    return output_file


# If a file is uploaded, process it and display a download link for the output file
if uploaded_file is not None:
    output_file = ms_table_creation_web(uploaded_file, year_start_value, nb_years_value, title_tab, header_color, unit_text, ms_text, source_text, note_text)
    st.download_button('Download processed file', output_file, file_name='processed_file.xlsx')
